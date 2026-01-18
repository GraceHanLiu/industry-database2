import pandas as pd
import logging
import os
from data_reader import DataReader
from data_processor import DataProcessor
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import time
from openpyxl.utils import get_column_letter
from typing import Dict
from contextlib import contextmanager
from datetime import datetime
import shutil

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class TimingStats:
    """用于记录各个步骤耗时的类"""
    def __init__(self):
        self.stats: Dict[str, float] = {}
        self.start_times: Dict[str, float] = {}
        
    @contextmanager
    def timing(self, step_name: str):
        """上下文管理器，用于记录某个步骤的耗时"""
        start_time = time.time()
        try:
            yield
        finally:
            duration = time.time() - start_time
            self.stats[step_name] = duration
            
    def get_total_time(self) -> float:
        """获取总耗时"""
        return sum(self.stats.values())
    
    def print_stats(self):
        """打印所有步骤的耗时统计"""
        print("\n=== 性能统计 ===")
        print(f"{'步骤':<30}{'耗时(秒)':<10}{'占比':<10}")
        print("-" * 50)
        total_time = self.get_total_time()
        for step, duration in self.stats.items():
            percentage = (duration / total_time) * 100 if total_time > 0 else 0
            print(f"{step:<30}{duration:>8.2f}s  {percentage:>6.1f}%")
        print("-" * 50)
        print(f"{'总计':<30}{total_time:>8.2f}s  100.0%")
        print("=" * 50)

def create_backup_file(original_file: str) -> str:
    """为目标文件创建带时间戳的副本，返回副本路径"""
    file_dir = os.path.dirname(original_file)
    file_name = os.path.basename(original_file)
    name, ext = os.path.splitext(file_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = os.path.join(file_dir, f"{name}_{timestamp}{ext}")
    shutil.copy2(original_file, backup_file)
    logger.info(f"已创建目标文件副本: {backup_file}")
    return backup_file

def transfer_data(source_file, target_file, source_sheet, target_sheet, indicator_code, 
                 target_start_row, target_start_col, backup_file=None, ascending=False, date_format='yyyy-mm-dd'):
    try:
        # 将字母列标识转换为数字
        if isinstance(target_start_col, str):
            target_start_col = column_letter_to_number(target_start_col)
            
        # 初始化计时器
        timing_stats = TimingStats()
        
        # 1. 创建或使用现有的副本文件
        if backup_file is None:
            with timing_stats.timing("创建目标文件副本"):
                backup_file = create_backup_file(target_file)
                print(f"\n数据将写入新副本: {backup_file}")
        else:
            print(f"\n数据将写入现有副本: {backup_file}")
        
        # 2. 初始化数据读取器
        with timing_stats.timing("初始化数据读取器"):
            reader = DataReader(source_file)
            processor = DataProcessor(reader)
        
        # 3. 读取指标数据
        with timing_stats.timing("读取指标数据"):
            print(f"\n读取指标数据: 工作表 '{source_sheet}', 指标 '{indicator_code}'")
            indicator_data = reader.read_indicator_data(source_sheet, indicator_code)
        
        # 4. 数据排序
        with timing_stats.timing("数据排序"):
            sorted_data = processor.sort_by_date(indicator_data, ascending=ascending)  # 使用参数控制排序
            df = sorted_data["data"]
        
        # 5. 打开副本文件
        with timing_stats.timing("打开目标文件副本"):
            wb_target = openpyxl.load_workbook(backup_file)
            ws_target = wb_target[target_sheet]
        
        # 6. 准备数据
        with timing_stats.timing("数据预处理"):
            data_to_write = []
            for _, row in df.iterrows():
                data_to_write.append([row['日期'], row[df.columns[-1]]])
        
        # 7. 写入数据
        with timing_stats.timing("数据写入"):
            for i, row_data in enumerate(data_to_write):
                row_num = target_start_row + i
                ws_target.cell(row=row_num, column=target_start_col, value=row_data[0])
                ws_target.cell(row=row_num, column=target_start_col + 1, value=row_data[1])
        
        # 8. 格式设置
        with timing_stats.timing("设置单元格格式"):
            date_col_letter = get_column_letter(target_start_col)
            date_range = f"{date_col_letter}{target_start_row}:{date_col_letter}{target_start_row + len(data_to_write) - 1}"
            for cell in ws_target[date_range]:
                cell[0].number_format = date_format
        
        # 9. 保存副本文件
        with timing_stats.timing("保存文件"):
            wb_target.save(backup_file)
        
        # 打印统计信息
        timing_stats.print_stats()
        
        logger.info(f"数据写入完成，总耗时: {timing_stats.get_total_time():.2f} 秒")
        for step, duration in timing_stats.stats.items():
            logger.info(f"{step} 耗时: {duration:.2f} 秒")
        
        print(f"\n成功写入 {len(data_to_write)} 行数据到副本: {backup_file}")
        print(f"原始目标文档未被修改: {target_file}")
        
        return backup_file, df['日期'].tolist()
        
    except Exception as e:
        logger.error(f"数据转移过程中出错: {str(e)}", exc_info=True)
        print(f"错误: {str(e)}")

def process_type1_task(source_file, target_file, source_sheet, target_sheet, indicators, start_row, start_col, backup_file, ascending=False, date_format='yyyy-mm-dd'):
    print(f"\n开始处理: 从'{source_sheet}'写入'{target_sheet}' (类型1)")
    # 将字母列标识转换为数字
    if isinstance(start_col, str):
        start_col_num = column_letter_to_number(start_col)
    else:
        start_col_num = start_col
    
    reference_dates = None
    
    for i, indicator in enumerate(indicators):
        if i == 0:
            backup_file, reference_dates = transfer_data(
                source_file=source_file,
                target_file=target_file,
                source_sheet=source_sheet,
                target_sheet=target_sheet,
                indicator_code=indicator,
                target_start_row=start_row,
                target_start_col=start_col,  
                backup_file=backup_file,
                ascending=ascending,
                date_format=date_format
            )
        else:
            reader = DataReader(source_file)
            current_data = reader.read_indicator_data(source_sheet, indicator)
            processor = DataProcessor(reader)
            sorted_data = processor.sort_by_date(current_data, ascending=ascending)
            df = sorted_data["data"]
        
        if df['日期'].tolist() == reference_dates:
            wb_target = openpyxl.load_workbook(backup_file)
            ws_target = wb_target[target_sheet]
            
            target_col = start_col_num + i + 1
            
            for row_idx, (_, row) in enumerate(df.iterrows()):
                target_row = start_row + row_idx
                value = row[df.columns[-1]]
                ws_target.cell(row=target_row, column=target_col, value=value)
            
            wb_target.save(backup_file)
            print(f"\n成功写入指标 {indicator} 的数据列")
        else:
            logger.warning(f"指标 {indicator} 的日期与第一个指标不匹配，跳过写入")
    
    return backup_file