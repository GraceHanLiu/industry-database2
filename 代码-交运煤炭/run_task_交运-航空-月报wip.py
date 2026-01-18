import pandas as pd
import logging
import os
from data_reader import DataReader
from data_processor import DataProcessor
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import time
from openpyxl.utils import get_column_letter
from typing import Dict
from contextlib import contextmanager
from datetime import datetime
import shutil

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class TimingStats:
    """用于记录各个步骤耗时的类"""
    def __init__(self):
        self.stats: Dict[str, float] = {}
        self.start_times: Dict[str, float] = {}
        
    @contextmanager
    def timing(self, step_name: str):
        """上下文管理器，用于记录某个步骤的耗时"""
        start_time = time.time()
        try:
            yield
        finally:
            duration = time.time() - start_time
            self.stats[step_name] = duration
            
    def get_total_time(self) -> float:
        """获取总耗时"""
        return sum(self.stats.values())
    
    def print_stats(self):
        """打印所有步骤的耗时统计"""
        print("\n=== 性能统计 ===")
        print(f"{'步骤':<30}{'耗时(秒)':<10}{'占比':<10}")
        print("-" * 50)
        total_time = self.get_total_time()
        for step, duration in self.stats.items():
            percentage = (duration / total_time) * 100 if total_time > 0 else 0
            print(f"{step:<30}{duration:>8.2f}s  {percentage:>6.1f}%")
        print("-" * 50)
        print(f"{'总计':<30}{total_time:>8.2f}s  100.0%")
        print("=" * 50)

def create_backup_file(original_file: str) -> str:
    """为目标文件创建带时间戳的副本，返回副本路径"""
    file_dir = os.path.dirname(original_file)
    file_name = os.path.basename(original_file)
    name, ext = os.path.splitext(file_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = os.path.join(file_dir, f"{name}_{timestamp}{ext}")
    shutil.copy2(original_file, backup_file)
    logger.info(f"已创建目标文件副本: {backup_file}")
    return backup_file

def transfer_data(source_file, target_file, source_sheet, target_sheet, indicator_code, 
                 target_start_row, target_start_col, backup_file=None, ascending=False, date_format='yyyy-mm-dd'):
    try:
        # 将字母列标识转换为数字
        if isinstance(target_start_col, str):
            target_start_col = column_letter_to_number(target_start_col)
            
        # 初始化计时器
        timing_stats = TimingStats()
        
        # 1. 创建或使用现有的副本文件
        if backup_file is None:
            with timing_stats.timing("创建目标文件副本"):
                backup_file = create_backup_file(target_file)
                print(f"\n数据将写入新副本: {backup_file}")
        else:
            print(f"\n数据将写入现有副本: {backup_file}")
        
        # 2. 初始化数据读取器
        with timing_stats.timing("初始化数据读取器"):
            reader = DataReader(source_file)
            processor = DataProcessor(reader)
        
        # 3. 读取指标数据
        with timing_stats.timing("读取指标数据"):
            print(f"\n读取指标数据: 工作表 '{source_sheet}', 指标 '{indicator_code}'")
            indicator_data = reader.read_indicator_data(source_sheet, indicator_code)
        
        # 4. 数据排序
        with timing_stats.timing("数据排序"):
            sorted_data = processor.sort_by_date(indicator_data, ascending=ascending)  # 使用参数控制排序
            df = sorted_data["data"]
        
        # 5. 打开副本文件
        with timing_stats.timing("打开目标文件副本"):
            wb_target = openpyxl.load_workbook(backup_file)
            ws_target = wb_target[target_sheet]
        
        # 6. 准备数据
        with timing_stats.timing("数据预处理"):
            data_to_write = []
            for _, row in df.iterrows():
                data_to_write.append([row['日期'], row[df.columns[-1]]])
        
        # 7. 写入数据
        with timing_stats.timing("数据写入"):
            for i, row_data in enumerate(data_to_write):
                row_num = target_start_row + i
                ws_target.cell(row=row_num, column=target_start_col, value=row_data[0])
                ws_target.cell(row=row_num, column=target_start_col + 1, value=row_data[1])
        
        # 8. 格式设置
        with timing_stats.timing("设置单元格格式"):
            date_col_letter = get_column_letter(target_start_col)
            date_range = f"{date_col_letter}{target_start_row}:{date_col_letter}{target_start_row + len(data_to_write) - 1}"
            for cell in ws_target[date_range]:
                cell[0].number_format = date_format
        
        # 9. 保存副本文件
        with timing_stats.timing("保存文件"):
            wb_target.save(backup_file)
        
        # 打印统计信息
        timing_stats.print_stats()
        
        logger.info(f"数据写入完成，总耗时: {timing_stats.get_total_time():.2f} 秒")
        for step, duration in timing_stats.stats.items():
            logger.info(f"{step} 耗时: {duration:.2f} 秒")
        
        print(f"\n成功写入 {len(data_to_write)} 行数据到副本: {backup_file}")
        print(f"原始目标文档未被修改: {target_file}")
        
        return backup_file, df['日期'].tolist()
        
    except Exception as e:
        logger.error(f"数据转移过程中出错: {str(e)}", exc_info=True)
        print(f"错误: {str(e)}")

def process_type1_task(source_file, target_file, source_sheet, target_sheet, indicators, start_row, start_col, 
                      backup_file, ascending=False, date_format='yyyy-mm-dd', indicator_data_cache=None):
    print(f"\n开始处理: 从'{source_sheet}'写入'{target_sheet}' (类型1)")
    if indicator_data_cache is None:
        indicator_data_cache = {}
    
    # 将字母列标识转换为数字
    if isinstance(start_col, str):
        start_col_num = column_letter_to_number(start_col)
    else:
        start_col_num = start_col
        
    reference_dates = None
    
    for i, indicator in enumerate(indicators):
        if i == 0:
            # 如果有缓存数据，使用缓存数据
            if indicator_data_cache and indicator in indicator_data_cache:
                df = indicator_data_cache[indicator]
                wb_target = openpyxl.load_workbook(backup_file)
                ws_target = wb_target[target_sheet]
                
                for row_idx, (_, row) in enumerate(df.iterrows()):
                    target_row = start_row + row_idx
                    ws_target.cell(row=target_row, column=start_col_num, value=row['日期'])
                    value = row[df.columns[-1]]
                    ws_target.cell(row=target_row, column=start_col_num + 1, value=value)
                
                wb_target.save(backup_file)
                reference_dates = df['日期'].tolist()
            else:
                backup_file, reference_dates = transfer_data(
                    source_file=source_file,
                    target_file=target_file,
                    source_sheet=source_sheet,
                    target_sheet=target_sheet,
                    indicator_code=indicator,
                    target_start_row=start_row,
                    target_start_col=start_col,  
                    backup_file=backup_file,
                    ascending=ascending,
                    date_format=date_format
                )
        else:
            # 使用缓存数据
            if indicator_data_cache and indicator in indicator_data_cache:
                df = indicator_data_cache[indicator]
            else:
                reader = DataReader(source_file)
                current_data = reader.read_indicator_data(source_sheet, indicator)
                processor = DataProcessor(reader)
                sorted_data = processor.sort_by_date(current_data, ascending=ascending)
                df = sorted_data["data"]
            
            if df['日期'].tolist() == reference_dates:
                wb_target = openpyxl.load_workbook(backup_file)
                ws_target = wb_target[target_sheet]
                
                target_col = start_col_num + i + 1
                
                for row_idx, (_, row) in enumerate(df.iterrows()):
                    target_row = start_row + row_idx
                    value = row[df.columns[-1]]
                    ws_target.cell(row=target_row, column=target_col, value=value)
                
                wb_target.save(backup_file)
                print(f"\n成功写入指标 {indicator} 的数据列")
            else:
                logger.warning(f"指标 {indicator} 的日期与第一个指标不匹配，跳过写入")
    
    return backup_file, indicator_data_cache

def process_type2_task(source_file, target_file, sheet_name, indicators, start_row, start_col, 
                      date_col, backup_file, ascending=False, date_format='yyyy-mm-dd'):
    """处理第二种类型的任务：使用已有日期列"""
    print(f"\n开始处理工作表: {sheet_name} (类型2)")
    
    # 将字母列标识转换为数字
    if isinstance(start_col, str):
        start_col_num = column_letter_to_number(start_col)
    else:
        start_col_num = start_col
        
    if isinstance(date_col, str):
        date_col_num = column_letter_to_number(date_col)
    else:
        date_col_num = date_col
    
    # 读取参考日期列
    wb_target = openpyxl.load_workbook(backup_file)
    ws_target = wb_target[sheet_name]
    reference_dates = []
    for row in range(start_row, ws_target.max_row + 1):
        date_value = ws_target.cell(row=row, column=date_col_num).value
        if date_value: 
            reference_dates.append(date_value)
        else:
            break

    # 处理每个指标
    for i, indicator in enumerate(indicators):
        reader = DataReader(source_file)
        current_data = reader.read_indicator_data(sheet_name, indicator)
        processor = DataProcessor(reader)
        sorted_data = processor.sort_by_date(current_data, ascending=ascending)
        df = sorted_data["data"]
        
        # 写入数据
        target_col = start_col_num + i
        for row_idx, (_, row) in enumerate(df.iterrows()):
            if row['日期'] in reference_dates:
                target_row = start_row + reference_dates.index(row['日期'])
                ws_target.cell(row=target_row, column=target_col, value=row[df.columns[-1]])
    
    wb_target.save(backup_file)
    print(f"\n成功写入工作表 {sheet_name} 的所有指标数据")
    
    return backup_file

# 在文件顶部导入部分下方添加两个辅助函数

def column_letter_to_number(column_letter):
    """将Excel列字母转换为数字索引（A->1, B->2, AA->27等）"""
    result = 0
    for char in column_letter:
        result = result * 26 + (ord(char.upper()) - ord('A') + 1)
    return result

def column_number_to_letter(column_number):
    """将数字索引转换为Excel列字母（1->A, 2->B, 27->AA等）"""
    result = ""
    while column_number > 0:
        column_number, remainder = divmod(column_number - 1, 26)
        result = chr(ord('A') + remainder) + result
    return result

def adjust_cell_values(backup_file, sheet_name, adjustments):
    """调整特定单元格的值和格式
    adjustments: [
        {
            'col': 'B',           # 列号（字母形式）
            'start_row': 2,       # 起始行
            'operation': '/',      # 操作类型
            'value': 10000,       # 操作数值
            'format': '0.00'      # 格式化字符串
        }
    ]
    """
    wb = openpyxl.load_workbook(backup_file)
    ws = wb[sheet_name]
    
    for adj in adjustments:
        col_num = column_letter_to_number(adj['col'])
        for row in range(adj['start_row'], ws.max_row + 1):
            cell = ws.cell(row=row, column=col_num)
            if cell.value is not None:
                if adj['operation'] == '/':
                    cell.value = cell.value / adj['value']
                elif adj['operation'] == '%':
                    cell.value = cell.value / 100
                cell.number_format = adj['format']
    
    wb.save(backup_file)
    print(f"\n已完成工作表 {sheet_name} 的数值调整")



def apply_formula(backup_file, sheet_name, config):
    """应用自定义公式到指定列
    Args:
        backup_file: Excel文件路径
        sheet_name: 工作表名称
        config: 公式配置，包含以下字段:
            {
                'target_col': 'C',  # 目标列（写入公式的列）
                'formula_type': 'ratio|yoy|mom|custom',  # 公式类型
                'params': {  # 根据formula_type不同而不同的参数
                    # 比值计算参数示例
                    'col1': 'A',  # 第一个列
                    'col2': 'B',  # 第二个列
                    # 同比/环比计算参数示例
                    'source_col': 'B',  # 源数据列
                    'periods': 12,  # 同比间隔(月)，环比为1
                    # 自定义公式参数示例
                    'formula_template': '={col1}{row}+{col2}{row}',  # 自定义公式模板
                    'cols': ['A', 'B']  # 公式中涉及的列
                },
                'start_row': 3,  # 开始行
                'end_row': None,  # 结束行，None表示自动检测
                'format': '0.00%'  # 单元格格式
            }
    """
    wb = openpyxl.load_workbook(backup_file)
    ws = wb[sheet_name]
    
    target_col = config['target_col']
    target_col_num = column_letter_to_number(target_col)
    formula_type = config['formula_type']
    start_row = config['start_row']
    format_str = config.get('format', 'General')
    
    # 确定结束行
    end_row = config.get('end_row')
    if end_row is None:
        # 自动检测结束行
        if formula_type in ['ratio','custom','custom2']:
            # 对于比值和自定义公式，检查所有相关列的最大有效行
            cols_to_check = []
            if formula_type == 'ratio':
                cols_to_check = [config['params']['col1'], config['params']['col2']]
            else:  # custom
                cols_to_check = config['params'].get('cols', [])
            
            end_row = start_row
            for col in cols_to_check:
                col_num = column_letter_to_number(col)
                for row in range(ws.max_row, start_row - 1, -1):
                    if ws.cell(row=row, column=col_num).value is not None:
                        end_row = max(end_row, row)
                        break
        elif formula_type in ['yoy', 'mom']:
            # 对于同比/环比，检查源数据列的最大有效行
            source_col = config['params']['source_col']
            source_col_num = column_letter_to_number(source_col)
            for row in range(ws.max_row, start_row - 1, -1):
                if ws.cell(row=row, column=source_col_num).value is not None:
                    end_row = row
                    break
    
    # 应用公式
    for row in range(start_row, end_row + 1):
        formula = ""
        
        if formula_type == 'ratio':
            # 比值计算: A/B
            col1 = config['params']['col1']
            col2 = config['params']['col2']
            formula = f"={col1}{row}/{col2}{row}"
        
        elif formula_type == 'yoy':
            # 同比计算: (B当前/B去年同期)-1
            source_col = config['params']['source_col']
            periods = config['params'].get('periods', 12)
            prev_row = row - periods
            if prev_row >= 1:  # 确保前一期的行号有效
                formula = f"={source_col}{row}/{source_col}{prev_row}-1"
        
        elif formula_type == 'mom':
            # 环比计算: (B当前/B上期)-1
            source_col = config['params']['source_col']
            if row > start_row:  # 确保有上一期数据
                formula = f"={source_col}{row}/{source_col}{row-1}-1"
        
        # elif formula_type == 'custom':
            # 自定义公式，使用模板替换
            # template = config['params']['formula_template']
            # formula = template.format(row=row, **{f"col{i+1}": col for i, col in enumerate(config['params'].get('cols', []))})
        elif formula_type == 'custom':
            template = config['params']['formula_template']
            prev_row = row - 12
            if prev_row >= 1:
                formula = template.format(row=row, prev_row=prev_row, **{f"col{i+1}": col for i, col in enumerate(config['params'].get('cols', []))})
        elif formula_type == 'custom2':
            # custom2: 当前行与64~75行循环相减
            template = config['params']['formula_template']
            cols = config['params'].get('cols', []) 
            if not cols: 
                continue
            col1 = cols[0]
            # 计算循环行号
            cycle_rows = list(range(64, 76))  # 64~75
            cycle_len = len(cycle_rows) 
            prev_row = cycle_rows[(row - start_row) % cycle_len]
            formula = template.format(row=row, prev_row=prev_row, col1=col1)

        # 写入公式
        if formula:
            cell = ws.cell(row=row, column=target_col_num)
            cell.value = formula
            cell.number_format = format_str
    
    wb.save(backup_file)
    print(f"\n已将公式应用到 {sheet_name} 工作表 {target_col} 列 (行 {start_row} 至 {end_row})")
    return backup_file


def write_airline_sheet(
    source_file, target_file, backup_file, indicator_col_map, 
    source_sheet, target_sheet, start_row=4, start_date="2014-01-01", date_format='yyyy-mm-dd'
):
    # 1. 读取所有指标数据
    reader = DataReader(source_file)
    processor = DataProcessor(reader)
    indicator_dfs = {}
    for indicator in indicator_col_map:
        ind_data = reader.read_indicator_data(source_sheet, indicator)
        df = ind_data["data"]
        # print(df['日期'])
        # 只保留start_date及以后的数据
        df = df[df['日期'] >= start_date]
        # 按时间正序
        df = df.sort_values('日期', ascending=True)
        indicator_dfs[indicator] = df.reset_index(drop=True)

    # 2. 打开目标文件
    wb = openpyxl.load_workbook(backup_file)
    ws = wb[target_sheet]

    # 3. 写入第一个指标的时间列和数据
    first_indicator = list(indicator_col_map.keys())[0]
    first_col = indicator_col_map[first_indicator]
    first_df = indicator_dfs[first_indicator]
    for i, row in first_df.iterrows():
        ws.cell(row=start_row + i, column=1, value=row['日期'])  # A列
        ws.cell(row=start_row + i, column=column_letter_to_number(first_col), value=row[first_df.columns[-1]])
        ws.cell(row=start_row + i, column=1).number_format = date_format

    # 4. 写入其他指标数据（不写时间列）
    for indicator, col in list(indicator_col_map.items())[1:]:
        df = indicator_dfs[indicator]
        # 对齐到第一个指标的日期
        date_list = first_df['日期'].tolist()
        value_map = dict(zip(df['日期'], df[df.columns[-1]]))
        for i, date in enumerate(date_list):
            value = value_map.get(date, None)
            ws.cell(row=start_row + i, column=column_letter_to_number(col), value=value)

    wb.save(backup_file)
    print(f"已完成'{target_sheet}'工作表的数据写入")

if __name__ == "__main__":
    try:
        # 文件路径
        source_file = 'D:\【交付审核】\_导出验收\Excel数据验收\交运\交通运输20250416_航空_航空月度数据_20250512070459.xlsx'
        target_file = "D:\【交付审核】\_导出验收\Excel数据验收\交运\交运航空目标文档.xlsx"
        
        # 创建一个副本
        backup_file = create_backup_file(target_file)
        print(f"\n已创建目标文件副本: {backup_file}")
        
        # 初始化指标数据缓存
        indicator_data_cache = {}

        # 处理"中国国航"工作表        
        indicator_col_map = {
            "AA000058863400": "B",
            "AA000058863500": "E",
            "AA000058863600": "H",
            "AA000058863700": "K",
            "AA000058869400": "N",
            "AA000058869500": "Q",
            "AA000058869600": "T",
            "AA000058869700": "W",
            "AA000058873400": "Z",
            "AA000058873500": "AC",
            "AA000058873600": "AF",
            "AA000058873700": "AI",
            "AA000058877400": "AL",
            "AA000058877500": "AO",
            "AA000058877600": "AR",
            "AA000058877700": "AU",
            "AA000064091900": "AX"
        }
        write_airline_sheet(
            source_file=source_file,
            target_file=target_file,
            backup_file=backup_file,
            indicator_col_map=indicator_col_map,
            source_sheet="中国国航",
            target_sheet="中国国航",
            start_row=4,
            start_date="2014-01-01",  # 可根据需要调整
            date_format='yyyy-mm'
        )

        # 衍生计算：同比上年
        yoy_cols = ["C", "F","I", "L", "O", "R", "U", "X", "AA", "AD", "AG", "AJ", "AY"]
        for col in yoy_cols:
            col_num = column_letter_to_number(col)
            source_col = column_number_to_letter(col_num - 1)
            config = {
                "target_col": col,
                "formula_type": "yoy",
                "params": {
                    "source_col": source_col,
                    "periods": 12
                },
                "start_row": 16,                
                "format": "0.00%"
            }
            apply_formula(
                backup_file=backup_file,
                sheet_name="中国国航", 
                config=config
            )

        
        yoy_diff_cols = ["AM", "AP", "AS", "AV"]
        for col in yoy_diff_cols:
            col_num = column_letter_to_number(col)
            source_col = column_number_to_letter(col_num - 1)
        
            config = {
                "target_col": col,
                "formula_type": "custom",
                "params": {
                    "formula_template": "={col1}{row}-{col1}{prev_row}",
                    "cols": [source_col]
                },
                "start_row": 16,
                "format": "0.00"
            }
            apply_formula(backup_file, "中国国航", config)

        # 衍生计算：同比19
        yoy19_cols = ["D", "G", "J", "M", "P", "S", "V", "Y", "AB", "AE", "AH", "AK", "AZ"]
        for col in yoy19_cols:
            col_num = column_letter_to_number(col)
            source_col = column_number_to_letter(col_num - 2)
            config = {
                "target_col": col,
                "formula_type": "custom2",
                "params": {
                    "formula_template": "={col1}{row}/{col1}{prev_row}-1",
                    "cols": [source_col]
                },
                "start_row": 112,
                "format": "0.00%"
            }
            apply_formula(
                backup_file=backup_file,
                sheet_name="中国国航",  # 如需应用于其他sheet请修改
                config=config
            )


        diff19_2_cols = ["AN", "AQ", "AT", "AW"]
        for col in diff19_2_cols:
            col_num = column_letter_to_number(col)
            source_col = column_number_to_letter(col_num - 2)
        
            config = {
                "target_col": col,
                "formula_type": "custom2",
                "params": {
                    "formula_template": "={col1}{row}-{col1}{prev_row}",
                    "cols": [source_col]
                },
                "start_row": 112,
                "format": "0.00"
            }
            apply_formula(backup_file, "中国国航", config)



        # # 处理"中国东航"工作表       、
        indicator_col_map_ceair = {
            "AA000058720500": "B",
            "AA000058720600": "E",
            "AA000058720700": "H",
            "AA000063550000": "K",
            "AA000058722600": "N",
            "AA000058722700": "Q",
            "AA000058722800": "T",
            "AA000063551300": "W",
            "AA000058723200": "Z",
            "AA000058723300": "AC",
            "AA000058723400": "AF",
            "AA000063552600": "AI",
            "AA000058721700": "AL",
            "AA000058721800": "AO",
            "AA000058721900": "AR",
            "AA000063553900": "AU",
            "AA000063567400": "AX"
        }
        write_airline_sheet(
            source_file=source_file,
            target_file=target_file,
            backup_file=backup_file,
            indicator_col_map=indicator_col_map_ceair,
            source_sheet="东方航空",
            target_sheet="中国东航",
            start_row=4,
            start_date="2014-01-01",  # 可根据需要调整
            date_format='yyyy-mm'
        )
      
        sheet_to_write = "中国东航"
        # 衍生计算：同比上年
        yoy_cols = ["C", "F","I", "L", "O", "R", "U", "X", "AA", "AD", "AG", "AJ", "AY"]
        for col in yoy_cols:
            col_num = column_letter_to_number(col)
            source_col = column_number_to_letter(col_num - 1)
            config = {
                "target_col": col,
                "formula_type": "yoy",
                "params": {
                    "source_col": source_col,
                    "periods": 12
                },
                "start_row": 16,                
                "format": "0.00%"
            }
            apply_formula(
                backup_file=backup_file,
                sheet_name= sheet_to_write, 
                config=config
            )

        
        yoy_diff_cols = ["AM", "AP", "AS", "AV"]
        for col in yoy_diff_cols:
            col_num = column_letter_to_number(col)
            source_col = column_number_to_letter(col_num - 1)
        
            config = {
                "target_col": col,
                "formula_type": "custom",
                "params": {
                    "formula_template": "={col1}{row}-{col1}{prev_row}",
                    "cols": [source_col]
                },
                "start_row": 16,
                "format": "0.00"
            }
            apply_formula(backup_file, sheet_to_write, config)

        # 衍生计算：同比19
        yoy19_cols = ["D", "G", "J", "M", "P", "S", "V", "Y", "AB", "AE", "AH", "AK", "AZ"]
        for col in yoy19_cols:
            col_num = column_letter_to_number(col)
            source_col = column_number_to_letter(col_num - 2)
            config = {
                "target_col": col,
                "formula_type": "custom2",
                "params": {
                    "formula_template": "={col1}{row}/{col1}{prev_row}-1",
                    "cols": [source_col]
                },
                "start_row": 112,
                "format": "0.00%"
            }
            apply_formula(
                backup_file=backup_file,
                sheet_name=sheet_to_write,  # 如需应用于其他sheet请修改
                config=config
            )


        diff19_2_cols = ["AN", "AQ", "AT", "AW"]
        for col in diff19_2_cols:
            col_num = column_letter_to_number(col)
            source_col = column_number_to_letter(col_num - 2)
        
            config = {
                "target_col": col,
                "formula_type": "custom2",
                "params": {
                    "formula_template": "={col1}{row}-{col1}{prev_row}",
                    "cols": [source_col]
                },
                "start_row": 112,
                "format": "0.00"
            }
            apply_formula(backup_file, sheet_to_write, config)


        print(f"\n！！！所有数据处理完成！最终文件保存为: {backup_file}")
        
    except Exception as e:
        logger.error(f"程序运行出错: {str(e)}", exc_info=True)
        print(f"\n程序运行出错: {str(e)}")