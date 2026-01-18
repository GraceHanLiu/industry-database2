import pandas as pd
import logging
import os
from data_reader import DataReader
from data_processor import DataProcessor
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import time
from openpyxl.utils import get_column_letter
from typing import Dict
from contextlib import contextmanager
from datetime import datetime
import shutil

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class TimingStats:
    """用于记录各个步骤耗时的类"""
    def __init__(self):
        self.stats: Dict[str, float] = {}
        self.start_times: Dict[str, float] = {}
        
    @contextmanager
    def timing(self, step_name: str):
        """上下文管理器，用于记录某个步骤的耗时"""
        start_time = time.time()
        try:
            yield
        finally:
            duration = time.time() - start_time
            self.stats[step_name] = duration
            
    def get_total_time(self) -> float:
        """获取总耗时"""
        return sum(self.stats.values())
    
    def print_stats(self):
        """打印所有步骤的耗时统计"""
        print("\n=== 性能统计 ===")
        print(f"{'步骤':<30}{'耗时(秒)':<10}{'占比':<10}")
        print("-" * 50)
        total_time = self.get_total_time()
        for step, duration in self.stats.items():
            percentage = (duration / total_time) * 100 if total_time > 0 else 0
            print(f"{step:<30}{duration:>8.2f}s  {percentage:>6.1f}%")
        print("-" * 50)
        print(f"{'总计':<30}{total_time:>8.2f}s  100.0%")
        print("=" * 50)

def create_backup_file(original_file: str) -> str:
    """为目标文件创建带时间戳的副本，返回副本路径"""
    file_dir = os.path.dirname(original_file)
    file_name = os.path.basename(original_file)
    name, ext = os.path.splitext(file_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = os.path.join(file_dir, f"{name}_{timestamp}{ext}")
    shutil.copy2(original_file, backup_file)
    logger.info(f"已创建目标文件副本: {backup_file}")
    return backup_file

def transfer_data(source_file, target_file, source_sheet, target_sheet, indicator_code, target_start_row, target_start_col, backup_file=None):
    """
    将源文件中的指定指标数据转移到目标文件的指定位置
    
    Parameters:
    source_file (str): 数据导出页文件路径
    target_file (str): 宏观数据库样例文件路径
    source_sheet (str): 源文件中的sheet名称
    target_sheet (str): 目标文件中的sheet名称
    indicator_code (str): 指标代码
    target_start_row (int): 目标文件中的起始行
    target_start_col (int): 目标文件中的起始列
    backup_file (str, optional): 已存在的副本文件路径
    """
    try:
        # 初始化计时器
        timing_stats = TimingStats()
        
        # 1. 创建或使用现有的副本文件
        if backup_file is None:
            with timing_stats.timing("创建目标文件副本"):
                backup_file = create_backup_file(target_file)
                print(f"\n数据将写入新副本: {backup_file}")
        else:
            print(f"\n数据将写入现有副本: {backup_file}")
        
        # 2. 初始化数据读取器
        with timing_stats.timing("初始化数据读取器"):
            reader = DataReader(source_file)
            processor = DataProcessor(reader)
        
        # 3. 读取指标数据
        with timing_stats.timing("读取指标数据"):
            print(f"\n读取指标数据: 工作表 '{source_sheet}', 指标 '{indicator_code}'")
            indicator_data = reader.read_indicator_data(source_sheet, indicator_code)
        
        # 4. 数据排序
        with timing_stats.timing("数据排序"):
            sorted_data = processor.sort_by_date(indicator_data, ascending=False) 
            df = sorted_data["data"]
        
        # 5. 打开副本文件
        with timing_stats.timing("打开目标文件副本"):
            wb_target = openpyxl.load_workbook(backup_file)
            ws_target = wb_target[target_sheet]
        
        # 6. 准备数据
        with timing_stats.timing("数据预处理"):
            data_to_write = []
            for _, row in df.iterrows():
                data_to_write.append([row['日期'], row[df.columns[-1]]])
        
        # 7. 写入数据
        with timing_stats.timing("数据写入"):
            for i, row_data in enumerate(data_to_write):
                row_num = target_start_row + i
                ws_target.cell(row=row_num, column=target_start_col, value=row_data[0])
                ws_target.cell(row=row_num, column=target_start_col + 1, value=row_data[1])
        
        # 8. 格式设置
        with timing_stats.timing("设置单元格格式"):
            date_col_letter = get_column_letter(target_start_col)
            date_range = f"{date_col_letter}{target_start_row}:{date_col_letter}{target_start_row + len(data_to_write) - 1}"
            for cell in ws_target[date_range]:
                cell[0].number_format = 'yyyy-mm-dd'
        
        # 9. 保存副本文件
        with timing_stats.timing("保存文件"):
            wb_target.save(backup_file)
        
        # 打印统计信息
        timing_stats.print_stats()
        
        logger.info(f"数据写入完成，总耗时: {timing_stats.get_total_time():.2f} 秒")
        for step, duration in timing_stats.stats.items():
            logger.info(f"{step} 耗时: {duration:.2f} 秒")
        
        print(f"\n成功写入 {len(data_to_write)} 行数据到副本: {backup_file}")
        print(f"原始目标文档未被修改: {target_file}")
        
    except Exception as e:
        logger.error(f"数据转移过程中出错: {str(e)}", exc_info=True)
        print(f"错误: {str(e)}")

if __name__ == "__main__":
    try:
        # 文件路径
        source_file = "D:\【交付审核】\_导出验收\Excel数据验收\数据插入法\煤炭\煤炭20250416_申万宏源煤炭行业数据库_20250429120906.xlsx"
        target_file = "D:\【交付审核】\_导出验收\Excel数据验收\数据插入法\煤炭\煤炭目标文档.xlsx"
        
        # 创建一个副本
        backup_file = create_backup_file(target_file)
        print(f"\n已创建目标文件副本: {backup_file}")
        
        # 第一个数据转移任务：美元汇率
        transfer_data(
            source_file=source_file,
            target_file=target_file,
            source_sheet="美元汇率",
            target_sheet="美元汇率",
            indicator_code="AA000085865500",
            target_start_row=12,
            target_start_col=3,
            backup_file=backup_file
        )
        
        # 第二个数据转移任务：布伦特原油价格
        transfer_data(
            source_file=source_file,
            target_file=target_file,
            source_sheet="布伦特原油价格",
            target_sheet="布伦特原油价格",
            indicator_code="AA000102805000",
            target_start_row=14,
            target_start_col=3,
            backup_file=backup_file
        )
        
        # 第三组数据转移任务：港口锚地船舶指标
        port_indicators = [
            "AA000061771600",
            "AA000061771700",
            "AA000061772600",
            "AA000061772700"
        ]
        
        # 存储第一个指标的日期数据，用于后续比对
        first_indicator_dates = None
        
        # 循环处理每个港口指标
        for i, indicator in enumerate(port_indicators):
            if i == 0:
                # 第一个指标：正常写入日期和数据
                transfer_data(
                    source_file=source_file,
                    target_file=target_file,
                    source_sheet="港口锚地船舶",
                    target_sheet="港口锚地船舶",
                    indicator_code=indicator,
                    target_start_row=13,
                    target_start_col=3,  # 从第3列开始写入日期和数据
                    backup_file=backup_file
                )
                
                # 读取第一个指标的数据，获取日期列
                reader = DataReader(source_file)
                first_data = reader.read_indicator_data("港口锚地船舶", indicator)
                processor = DataProcessor(reader)
                sorted_data = processor.sort_by_date(first_data, ascending=False)
                first_indicator_dates = sorted_data["data"]['日期'].tolist()
                
            else:
                # 后续指标：只写入数据列，紧贴前一个指标
                reader = DataReader(source_file)
                current_data = reader.read_indicator_data("港口锚地船舶", indicator)
                processor = DataProcessor(reader)
                sorted_data = processor.sort_by_date(current_data, ascending=False)
                df = sorted_data["data"]
                
                # 确保日期匹配
                if df['日期'].tolist() == first_indicator_dates:
                    # 打开副本文件
                    wb_target = openpyxl.load_workbook(backup_file)
                    ws_target = wb_target["港口锚地船舶"]
                    
                    # 只写入数据列，计算正确的列位置
                    # 第一个指标占用列3(日期)和列4(数据)，后续指标从列5开始依次写入
                    target_col = 4 + i  # 修改这里：4是第一个指标的数据列，后续指标依次加1
                    
                    # 写入数据
                    for row_idx, (_, row) in enumerate(df.iterrows()):
                        target_row = 13 + row_idx
                        ws_target.cell(row=target_row, column=target_col, value=row[df.columns[-1]])
                    
                    # 保存文件
                    wb_target.save(backup_file)
                    print(f"\n成功写入指标 {indicator} 的数据列")
                else:
                    logger.warning(f"指标 {indicator} 的日期与第一个指标不匹配，跳过写入")
        
    except Exception as e:
        logger.error(f"程序运行出错: {str(e)}", exc_info=True)
        print(f"\n程序运行出错: {str(e)}")