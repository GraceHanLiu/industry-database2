import pandas as pd
import logging
import os
from data_reader import DataReader
from data_processor import DataProcessor
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import time
from openpyxl.utils import get_column_letter
from typing import Dict
from contextlib import contextmanager
from datetime import datetime
import shutil

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class TimingStats:
    """用于记录各个步骤耗时的类"""
    def __init__(self):
        self.stats: Dict[str, float] = {}
        self.start_times: Dict[str, float] = {}
        
    @contextmanager
    def timing(self, step_name: str):
        """上下文管理器，用于记录某个步骤的耗时"""
        start_time = time.time()
        try:
            yield
        finally:
            duration = time.time() - start_time
            self.stats[step_name] = duration
            
    def get_total_time(self) -> float:
        """获取总耗时"""
        return sum(self.stats.values())
    
    def print_stats(self):
        """打印所有步骤的耗时统计"""
        print("\n=== 性能统计 ===")
        print(f"{'步骤':<30}{'耗时(秒)':<10}{'占比':<10}")
        print("-" * 50)
        total_time = self.get_total_time()
        for step, duration in self.stats.items():
            percentage = (duration / total_time) * 100 if total_time > 0 else 0
            print(f"{step:<30}{duration:>8.2f}s  {percentage:>6.1f}%")
        print("-" * 50)
        print(f"{'总计':<30}{total_time:>8.2f}s  100.0%")
        print("=" * 50)

def create_backup_file(original_file: str) -> str:
    """为目标文件创建带时间戳的副本，返回副本路径"""
    file_dir = os.path.dirname(original_file)
    file_name = os.path.basename(original_file)
    name, ext = os.path.splitext(file_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = os.path.join(file_dir, f"{name}_{timestamp}{ext}")
    shutil.copy2(original_file, backup_file)
    logger.info(f"已创建目标文件副本: {backup_file}")
    return backup_file

def transfer_data(source_file, target_file, source_sheet, target_sheet, indicator_code, 
                 target_start_row, target_start_col, backup_file=None, ascending=False, date_format='yyyy-mm-dd'):
    try:
        # 将字母列标识转换为数字
        if isinstance(target_start_col, str):
            target_start_col = column_letter_to_number(target_start_col)
            
        # 初始化计时器
        timing_stats = TimingStats()
        
        # 1. 创建或使用现有的副本文件
        if backup_file is None:
            with timing_stats.timing("创建目标文件副本"):
                backup_file = create_backup_file(target_file)
                print(f"\n数据将写入新副本: {backup_file}")
        else:
            print(f"\n数据将写入现有副本: {backup_file}")
        
        # 2. 初始化数据读取器
        with timing_stats.timing("初始化数据读取器"):
            reader = DataReader(source_file)
            processor = DataProcessor(reader)
        
        # 3. 读取指标数据
        with timing_stats.timing("读取指标数据"):
            print(f"\n读取指标数据: 工作表 '{source_sheet}', 指标 '{indicator_code}'")
            indicator_data = reader.read_indicator_data(source_sheet, indicator_code)
        
        # 4. 数据排序
        with timing_stats.timing("数据排序"):
            sorted_data = processor.sort_by_date(indicator_data, ascending=ascending)  # 使用参数控制排序
            df = sorted_data["data"]
        
        # 5. 打开副本文件
        with timing_stats.timing("打开目标文件副本"):
            wb_target = openpyxl.load_workbook(backup_file)
            ws_target = wb_target[target_sheet]
        
        # 6. 准备数据
        with timing_stats.timing("数据预处理"):
            data_to_write = []
            for _, row in df.iterrows():
                data_to_write.append([row['日期'], row[df.columns[-1]]])
        
        # 7. 写入数据
        with timing_stats.timing("数据写入"):
            for i, row_data in enumerate(data_to_write):
                row_num = target_start_row + i
                ws_target.cell(row=row_num, column=target_start_col, value=row_data[0])
                ws_target.cell(row=row_num, column=target_start_col + 1, value=row_data[1])
        
        # 8. 格式设置
        with timing_stats.timing("设置单元格格式"):
            date_col_letter = get_column_letter(target_start_col)
            date_range = f"{date_col_letter}{target_start_row}:{date_col_letter}{target_start_row + len(data_to_write) - 1}"
            for cell in ws_target[date_range]:
                cell[0].number_format = date_format
        
        # 9. 保存副本文件
        with timing_stats.timing("保存文件"):
            wb_target.save(backup_file)
        
        # 打印统计信息
        timing_stats.print_stats()
        
        logger.info(f"数据写入完成，总耗时: {timing_stats.get_total_time():.2f} 秒")
        for step, duration in timing_stats.stats.items():
            logger.info(f"{step} 耗时: {duration:.2f} 秒")
        
        print(f"\n成功写入 {len(data_to_write)} 行数据到副本: {backup_file}")
        print(f"原始目标文档未被修改: {target_file}")
        
        return backup_file, df['日期'].tolist()
        
    except Exception as e:
        logger.error(f"数据转移过程中出错: {str(e)}", exc_info=True)
        print(f"错误: {str(e)}")

def process_type1_task(source_file, target_file, source_sheet, target_sheet, indicators, start_row, start_col, 
                      backup_file, ascending=False, date_format='yyyy-mm-dd', indicator_data_cache=None):
    print(f"\n开始处理: 从'{source_sheet}'写入'{target_sheet}' (类型1)")
    if indicator_data_cache is None:
        indicator_data_cache = {}
    
    # 将字母列标识转换为数字
    if isinstance(start_col, str):
        start_col_num = column_letter_to_number(start_col)
    else:
        start_col_num = start_col
        
    reference_dates = None
    
    for i, indicator in enumerate(indicators):
        if i == 0:
            # 如果有缓存数据，使用缓存数据
            if indicator_data_cache and indicator in indicator_data_cache:
                df = indicator_data_cache[indicator]
                wb_target = openpyxl.load_workbook(backup_file)
                ws_target = wb_target[target_sheet]
                
                for row_idx, (_, row) in enumerate(df.iterrows()):
                    target_row = start_row + row_idx
                    ws_target.cell(row=target_row, column=start_col_num, value=row['日期'])
                    value = row[df.columns[-1]]
                    ws_target.cell(row=target_row, column=start_col_num + 1, value=value)
                
                wb_target.save(backup_file)
                reference_dates = df['日期'].tolist()
            else:
                backup_file, reference_dates = transfer_data(
                    source_file=source_file,
                    target_file=target_file,
                    source_sheet=source_sheet,
                    target_sheet=target_sheet,
                    indicator_code=indicator,
                    target_start_row=start_row,
                    target_start_col=start_col,  
                    backup_file=backup_file,
                    ascending=ascending,
                    date_format=date_format
                )
        else:
            # 使用缓存数据
            if indicator_data_cache and indicator in indicator_data_cache:
                df = indicator_data_cache[indicator]
            else:
                reader = DataReader(source_file)
                current_data = reader.read_indicator_data(source_sheet, indicator)
                processor = DataProcessor(reader)
                sorted_data = processor.sort_by_date(current_data, ascending=ascending)
                df = sorted_data["data"]
            
            if df['日期'].tolist() == reference_dates:
                wb_target = openpyxl.load_workbook(backup_file)
                ws_target = wb_target[target_sheet]
                
                target_col = start_col_num + i + 1
                
                for row_idx, (_, row) in enumerate(df.iterrows()):
                    target_row = start_row + row_idx
                    value = row[df.columns[-1]]
                    ws_target.cell(row=target_row, column=target_col, value=value)
                
                wb_target.save(backup_file)
                print(f"\n成功写入指标 {indicator} 的数据列")
            else:
                logger.warning(f"指标 {indicator} 的日期与第一个指标不匹配，跳过写入")
    
    return backup_file, indicator_data_cache

def process_type2_task(source_file, target_file, sheet_name, indicators, start_row, start_col, 
                      date_col, backup_file, ascending=False, date_format='yyyy-mm-dd'):
    """处理第二种类型的任务：使用已有日期列"""
    print(f"\n开始处理工作表: {sheet_name} (类型2)")
    
    # 将字母列标识转换为数字
    if isinstance(start_col, str):
        start_col_num = column_letter_to_number(start_col)
    else:
        start_col_num = start_col
        
    if isinstance(date_col, str):
        date_col_num = column_letter_to_number(date_col)
    else:
        date_col_num = date_col
    
    # 读取参考日期列
    wb_target = openpyxl.load_workbook(backup_file)
    ws_target = wb_target[sheet_name]
    reference_dates = []
    for row in range(start_row, ws_target.max_row + 1):
        date_value = ws_target.cell(row=row, column=date_col_num).value
        if date_value: 
            reference_dates.append(date_value)
        else:
            break

    # 处理每个指标
    for i, indicator in enumerate(indicators):
        reader = DataReader(source_file)
        current_data = reader.read_indicator_data(sheet_name, indicator)
        processor = DataProcessor(reader)
        sorted_data = processor.sort_by_date(current_data, ascending=ascending)
        df = sorted_data["data"]
        
        # 写入数据
        target_col = start_col_num + i
        for row_idx, (_, row) in enumerate(df.iterrows()):
            if row['日期'] in reference_dates:
                target_row = start_row + reference_dates.index(row['日期'])
                ws_target.cell(row=target_row, column=target_col, value=row[df.columns[-1]])
    
    wb_target.save(backup_file)
    print(f"\n成功写入工作表 {sheet_name} 的所有指标数据")
    
    return backup_file

# 在文件顶部导入部分下方添加两个辅助函数

def column_letter_to_number(column_letter):
    """将Excel列字母转换为数字索引（A->1, B->2, AA->27等）"""
    result = 0
    for char in column_letter:
        result = result * 26 + (ord(char.upper()) - ord('A') + 1)
    return result

def column_number_to_letter(column_number):
    """将数字索引转换为Excel列字母（1->A, 2->B, 27->AA等）"""
    result = ""
    while column_number > 0:
        column_number, remainder = divmod(column_number - 1, 26)
        result = chr(ord('A') + remainder) + result
    return result

def adjust_cell_values(backup_file, sheet_name, adjustments):
    """调整特定单元格的值和格式
    adjustments: [
        {
            'col': 'B',           # 列号（字母形式）
            'start_row': 2,       # 起始行
            'operation': '/',      # 操作类型
            'value': 10000,       # 操作数值
            'format': '0.00'      # 格式化字符串
        }
    ]
    """
    wb = openpyxl.load_workbook(backup_file)
    ws = wb[sheet_name]
    
    for adj in adjustments:
        col_num = column_letter_to_number(adj['col'])
        for row in range(adj['start_row'], ws.max_row + 1):
            cell = ws.cell(row=row, column=col_num)
            if cell.value is not None:
                if adj['operation'] == '/':
                    cell.value = cell.value / adj['value']
                elif adj['operation'] == '%':
                    cell.value = cell.value / 100
                elif adj['operation'] == '*':
                    cell.value = cell.value * adj['value']
                cell.number_format = adj['format']
    
    wb.save(backup_file)
    print(f"\n已完成工作表 {sheet_name} 的数值调整")



def apply_formula(backup_file, sheet_name, config):
    """应用自定义公式到指定列
    Args:
        backup_file: Excel文件路径
        sheet_name: 工作表名称
        config: 公式配置，包含以下字段:
            {
                'target_col': 'C',  # 目标列（写入公式的列）
                'formula_type': 'ratio|yoy|mom|custom',  # 公式类型
                'params': {  # 根据formula_type不同而不同的参数
                    # 比值计算参数示例
                    'col1': 'A',  # 第一个列
                    'col2': 'B',  # 第二个列
                    # 同比/环比计算参数示例
                    'source_col': 'B',  # 源数据列
                    'periods': 12,  # 同比间隔(月)，环比为1
                    # 自定义公式参数示例
                    'formula_template': '={col1}{row}+{col2}{row}',  # 自定义公式模板
                    'cols': ['A', 'B']  # 公式中涉及的列
                },
                'start_row': 3,  # 开始行
                'end_row': None,  # 结束行，None表示自动检测
                'format': '0.00%'  # 单元格格式
            }
    """
    wb = openpyxl.load_workbook(backup_file)
    ws = wb[sheet_name]
    
    target_col = config['target_col']
    target_col_num = column_letter_to_number(target_col)
    formula_type = config['formula_type']
    start_row = config['start_row']
    format_str = config.get('format', 'General')
    
    # 确定结束行
    end_row = config.get('end_row')
    if end_row is None:
        # 自动检测结束行
        if formula_type in ['ratio', 'custom']:
            # 对于比值和自定义公式，检查所有相关列的最大有效行
            cols_to_check = []
            if formula_type == 'ratio':
                cols_to_check = [config['params']['col1'], config['params']['col2']]
            else:  # custom
                cols_to_check = config['params'].get('cols', [])
            
            end_row = start_row
            for col in cols_to_check:
                col_num = column_letter_to_number(col)
                for row in range(ws.max_row, start_row - 1, -1):
                    if ws.cell(row=row, column=col_num).value is not None:
                        end_row = max(end_row, row)
                        break
        elif formula_type in ['yoy', 'mom']:
            # 对于同比/环比，检查源数据列的最大有效行
            source_col = config['params']['source_col']
            source_col_num = column_letter_to_number(source_col)
            for row in range(ws.max_row, start_row - 1, -1):
                if ws.cell(row=row, column=source_col_num).value is not None:
                    end_row = row
                    break
    
    # 应用公式
    for row in range(start_row, end_row + 1):
        formula = ""
        
        if formula_type == 'ratio':
            # 比值计算: A/B
            col1 = config['params']['col1']
            col2 = config['params']['col2']
            formula = f"={col1}{row}/{col2}{row}"
        
        elif formula_type == 'yoy':
            # 同比计算: (B当前/B去年同期)-1
            source_col = config['params']['source_col']
            periods = config['params'].get('periods', 12)
            prev_row = row - periods
            if prev_row >= 1:  # 确保前一期的行号有效
                formula = f"={source_col}{row}/{source_col}{prev_row}-1"
        
        elif formula_type == 'mom':
            # 环比计算: (B当前/B上期)-1
            source_col = config['params']['source_col']
            if row > start_row:  # 确保有上一期数据
                formula = f"={source_col}{row}/{source_col}{row-1}-1"
        
        elif formula_type == 'custom':
            # 自定义公式，使用模板替换
            template = config['params']['formula_template']
            formula = template.format(row=row, **{f"col{i+1}": col for i, col in enumerate(config['params'].get('cols', []))})
        
        # 写入公式
        if formula:
            cell = ws.cell(row=row, column=target_col_num)
            cell.value = formula
            cell.number_format = format_str
    
    wb.save(backup_file)
    print(f"\n已将公式应用到 {sheet_name} 工作表 {target_col} 列 (行 {start_row} 至 {end_row})")
    return backup_file

def write_specified_dates_to_dashboard(backup_file, dashboard_sheet, dashboard_col, dashboard_start_row, date_list):
    wb = openpyxl.load_workbook(backup_file)
    ws_dashboard = wb[dashboard_sheet]
    for i, date in enumerate(date_list):
        ws_dashboard.cell(row=dashboard_start_row + i, column=column_letter_to_number(dashboard_col), value=date)
    wb.save(backup_file)
    print(f"已将{len(date_list)}个指定时间点写入到【数据看板】sheet的{dashboard_col}{dashboard_start_row}起")



if __name__ == "__main__":
    try:
        # 文件路径
        source_file = 'D:\【交付审核】\_导出验收\Excel数据验收\交运\交通运输20250416_公路铁路_公路铁路物流数据_20250512072459.xlsx'
        target_file = "D:\【交付审核】\_导出验收\Excel数据验收\交运\交运公路目标文档.xlsx"
        
        # 创建一个副本
        backup_file = create_backup_file(target_file)
        print(f"\n已创建目标文件副本: {backup_file}")
        
        # 初始化指标数据缓存
        indicator_data_cache = {}

        # 处理"全国中心城市客运量"工作表
        sheet_name = "全国中心城市客运量"
        source_sheet = "全国中心城市客运量"
        indicators = [
            'AA000065670000',
            'AA000065667100',
            'AA000065672900'
        ]
        
        backup_file, indicator_data_cache = process_type1_task(
            source_file=source_file,
            target_file=target_file,
            source_sheet=source_sheet,
            target_sheet=sheet_name,
            indicators=indicators,
            start_row=3,  # 从A3开始写入
            start_col='A',  # 从A列开始
            backup_file=backup_file,
            ascending=False,  # 按时间倒序写入
            date_format='yyyy-mm-dd',
            indicator_data_cache=indicator_data_cache
        )
        

        
        # 处理"铁路运输"工作表
        sheet_name = "铁路运输"
        source_sheet = "铁路运输"
        indicators = [
            'AA000083792400',  # 除以10000
            'AA000084987200',
            'AA000083792200',  # 除以10000
            'AA000083792300',
            'AA000083789400',  # 除以100000000
            'AA000084986200',
            'AA000083789200',  # 除以100000000
            'AA000083789300',
            'AA000083790900',  # 除以10000
            'AA000084986700',
            'AA000083790700',  # 除以10000
            'AA000083790800',
            'AA000083787900',  # 除以100000000
            'AA000084985700',
            'AA000083787700',  # 除以100000000
            'AA000083787800'
        ]
        
        # 定义需要调整的指标及其调整因子
        adjustments = {
            'AA000083792400': 10000,
            'AA000083792200': 10000,
            'AA000083790900': 10000,
            'AA000083790700': 10000,
            'AA000083789400': 100000000,
            'AA000083789200': 100000000,
            'AA000083787900': 100000000,
            'AA000083787700': 100000000
        }
        
        # 读取所有指标数据并筛选日期
        reader = DataReader(source_file)
        processor = DataProcessor(reader)
        
        # 存储筛选后的所有指标数据
        filtered_data = []
        for indicator in indicators:
            current_data = reader.read_indicator_data(source_sheet, indicator)
            sorted_data = processor.sort_by_date(current_data, ascending=False)  # 按时间倒序
            df = sorted_data["data"]
            # 将0值替换为None（这样Excel中会显示为空）
            df[df.columns[-1]] = df[df.columns[-1]].replace(0, None)
            
            # 如果指标需要调整，进行数值调整
            if indicator in adjustments:
                df[df.columns[-1]] = df[df.columns[-1]].apply(
                    lambda x: round(x / adjustments[indicator], 2) if x is not None else None
                )
            
            filtered_data.append(df)
        
        # 打开工作簿并获取目标工作表
        wb = openpyxl.load_workbook(backup_file)
        ws = wb[sheet_name]
        
        # 写入数据
        start_row = 3
        start_col = 'A'  # 时间列
        start_col_num = column_letter_to_number(start_col)
        
        # 获取第一个指标的数据作为参考
        reference_df = filtered_data[0]
        
        # 写入时间列和所有指标数据
        for row_idx, (_, row) in enumerate(reference_df.iterrows()):
            current_row = start_row + row_idx
            
            # 写入日期
            ws.cell(row=current_row, column=start_col_num, value=row['日期'])
            ws.cell(row=current_row, column=start_col_num).number_format = 'yyyy-mm-dd'
            
            # 写入所有指标数据
            for i, df in enumerate(filtered_data):
                value = df.loc[df['日期'] == row['日期'], df.columns[-1]].iloc[0]
                cell = ws.cell(row=current_row, column=start_col_num + i + 1, value=value)
                
                # 如果是需要调整的指标，设置数字格式为保留2位小数
                if indicators[i] in adjustments:
                    cell.number_format = '0.00'                 
        
        # 保存文件
        wb.save(backup_file)
        print(f"\n成功处理工作表: {sheet_name}")
                
        # 处理"铁路投资"工作表
        source_sheet = "铁路投资"
        target_sheet = "铁路投资"
        indicators = [
            'AA000065091700',  # 需乘以10000
            'AA000065336000'
        ]        
        backup_file, indicator_data_cache = process_type1_task(
            source_file=source_file,
            target_file=target_file,
            source_sheet=source_sheet,
            target_sheet=target_sheet,
            indicators=indicators,
            start_row=3,
            start_col='A',
            backup_file=backup_file,
            ascending=False,
            date_format='yyyy-mm-dd',
            indicator_data_cache=indicator_data_cache
        )
        
        adjustments = [
            {
                'col': 'B',           # AA000065091700写入在B列
                'start_row': 3,
                'operation': '*',     # 乘法操作
                'value': 10000,
                'format': '0.00'
            }
        ]
                
        adjust_cell_values(backup_file, target_sheet, adjustments)
        
        print(f"\n成功将铁路投资数据写入到工作表: {target_sheet}")


        # # 处理"主要省份公路累计货运量及同比"工作表
        sheet_name = "主要省份公路累计货运量及同比"
        source_sheet = "主要省份公路累计货运量及同比"
        indicators = [
            'AA000084622500', 'AA000084622400', 'AA000084617600', 'AA000084617500', 'AA000084617400',
            'AA000084617300', 'AA000084617200', 'AA000084617100', 'AA000084617000', 'AA000084616900',
            'AA000084616800', 'AA000084616700', 'AA000084616600', 'AA000084616500', 'AA000084616400',
            'AA000084616300', 'AA000084616200', 'AA000084616100', 'AA000084616000', 'AA000084615900',
            'AA000084615800', 'AA000084615700', 'AA000084615600', 'AA000084615500', 'AA000084615400',
            'AA000084615300', 'AA000084615200', 'AA000084615100', 'AA000084615000', 'AA000084614900',
            'AA000084614800'
        ]
        
        # 1. 先写入数据
        backup_file, indicator_data_cache = process_type1_task(
            source_file=source_file,
            target_file=target_file,
            source_sheet=source_sheet,
            target_sheet=sheet_name,
            indicators=indicators,
            start_row=4,  # 从A4开始写入
            start_col='A',  # 从A列开始
            backup_file=backup_file,
            ascending=False,  # 按时间倒序写入
            date_format='yyyy-mm-dd',
            indicator_data_cache=indicator_data_cache
        )
        
        # 2. 再批量调整所有数据列（B列及之后）
        adjustments = [
            {
                'col': chr(ord('B') + i),  # B、C、D、...
                'start_row': 4,
                'operation': '/',
                'value': 10000,
                'format': '0.00'
            }
            for i in range(len(indicators))
        ]
        adjust_cell_values(backup_file, sheet_name, adjustments)

        print(f"\n成功将数据写入到工作表: {target_sheet}")
    
        print(f"\n！！！所有数据处理完成！最终文件保存为: {backup_file}")
        
    except Exception as e:
        logger.error(f"程序运行出错: {str(e)}", exc_info=True)
        print(f"\n程序运行出错: {str(e)}")